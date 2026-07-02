# Copyright (c) 2026 Ravenkey LLC. All rights reserved.
"""Extract text from PDF, DOCX, XLSX, and email files."""

from __future__ import annotations

import csv
import email
import io
import os
import tempfile
from pathlib import Path


def extract_text(file_path: str) -> str:
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":
        return _extract_pdf(file_path)
    elif ext == ".docx":
        return _extract_docx(file_path)
    elif ext == ".xlsx":
        return _extract_xlsx(file_path)
    elif ext == ".csv":
        return _extract_csv(file_path)
    elif ext in (".eml", ".msg"):
        return _extract_email(file_path)
    elif ext in (".txt", ".md", ".html", ".htm"):
        return Path(file_path).read_text(encoding="utf-8", errors="replace")
    else:
        return ""


def _extract_pdf(file_path: str) -> str:
    from pypdf import PdfReader

    reader = PdfReader(file_path)
    parts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            parts.append(text)
    return "\n\n".join(parts)


def _extract_docx(file_path: str) -> str:
    from docx import Document

    doc = Document(file_path)
    parts = []
    for para in doc.paragraphs:
        if para.text.strip():
            parts.append(para.text)
    return "\n\n".join(parts)


def _extract_xlsx(file_path: str) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"[Sheet: {sheet_name}]")
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join(str(c) if c is not None else "" for c in row)
            if row_text.strip():
                parts.append(row_text)
    return "\n".join(parts)


def _extract_csv(file_path: str) -> str:
    parts = []
    with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            parts.append("\t".join(row))
    return "\n".join(parts)


def _extract_email(file_path: str) -> str:
    with open(file_path, "rb") as f:
        raw = f.read()
    try:
        msg = email.message_from_bytes(raw)
    except Exception:
        msg = email.message_from_string(raw.decode("utf-8", errors="replace"))
    parts = []
    parts.append(f"Subject: {msg.get('Subject', '')}")
    parts.append(f"From: {msg.get('From', '')}")
    parts.append(f"Date: {msg.get('Date', '')}")
    parts.append("")
    body = _get_email_body(msg)
    if body:
        parts.append(body)
    return "\n".join(parts)


def _get_email_body(msg) -> str:
    if msg.is_multipart():
        text_parts = []
        for part in msg.walk():
            ctype = part.get_content_type()
            if ctype == "text/plain":
                try:
                    text_parts.append(part.get_payload(decode=True).decode("utf-8", errors="replace"))
                except Exception:
                    text_parts.append(str(part.get_payload(decode=True)))
            elif ctype == "text/html" and not text_parts:
                try:
                    text_parts.append(part.get_payload(decode=True).decode("utf-8", errors="replace"))
                except Exception:
                    text_parts.append(str(part.get_payload(decode=True)))
        return "\n".join(text_parts)
    else:
        try:
            return msg.get_payload(decode=True).decode("utf-8", errors="replace")
        except Exception:
            return str(msg.get_payload(decode=True))
